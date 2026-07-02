import csv
import io
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.contrib.auth.models import User
from prediction.models import AuditLog


@login_required
def audit_log_list(request):
    audit_logs = AuditLog.objects.all()

    action_filter = request.GET.get('action', '')
    module_filter = request.GET.get('module', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    user_filter = request.GET.get('user', '')

    if action_filter:
        audit_logs = audit_logs.filter(action=action_filter)
    if module_filter:
        audit_logs = audit_logs.filter(module=module_filter)
    if date_from:
        audit_logs = audit_logs.filter(created_at__date__gte=date_from)
    if date_to:
        audit_logs = audit_logs.filter(created_at__date__lte=date_to)
    if user_filter:
        audit_logs = audit_logs.filter(user__username__icontains=user_filter)

    audit_logs = audit_logs.order_by('-created_at')

    paginator = Paginator(audit_logs, 15)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    modules = AuditLog.objects.values_list('module', flat=True).distinct().order_by('module')
    users = User.objects.filter(id__in=AuditLog.objects.values('user_id').distinct()).order_by('username')

    context = {
        'page_obj': page_obj,
        'audit_logs': page_obj,
        'action_filter': action_filter,
        'module_filter': module_filter,
        'date_from': date_from,
        'date_to': date_to,
        'user_filter': user_filter,
        'action_choices': AuditLog.ACTION_CHOICES,
        'modules': modules,
        'users': users,
        'page_title': '审计日志',
        'page_subtitle': '系统操作审计追踪',
    }
    return render(request, 'audit_log_list.html', context)


@login_required
def audit_log_detail(request, pk):
    audit_log = get_object_or_404(AuditLog, pk=pk)
    context = {
        'audit_log': audit_log,
        'page_title': '审计日志详情',
        'page_subtitle': f'日志 #{audit_log.pk}',
    }
    return render(request, 'audit_log_detail.html', context)


@login_required
def audit_log_export(request):
    audit_logs = AuditLog.objects.all()

    action_filter = request.GET.get('action', '')
    module_filter = request.GET.get('module', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    user_filter = request.GET.get('user', '')

    if action_filter:
        audit_logs = audit_logs.filter(action=action_filter)
    if module_filter:
        audit_logs = audit_logs.filter(module=module_filter)
    if date_from:
        audit_logs = audit_logs.filter(created_at__date__gte=date_from)
    if date_to:
        audit_logs = audit_logs.filter(created_at__date__lte=date_to)
    if user_filter:
        audit_logs = audit_logs.filter(user__username__icontains=user_filter)

    audit_logs = audit_logs.order_by('-created_at')

    export_format = request.GET.get('format', 'csv')

    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="audit_logs.csv"'
        response.write('\ufeff')

        writer = csv.writer(response)
        writer.writerow(['操作时间', '操作用户', '操作类型', '模块', '操作对象', '操作详情', 'IP地址', '用户代理'])

        action_display = dict(AuditLog.ACTION_CHOICES)

        for log in audit_logs:
            writer.writerow([
                log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                log.user.username if log.user else '匿名',
                action_display.get(log.action, '其他操作'),
                log.module,
                log.target or '',
                log.detail or '',
                log.ip_address or '',
                log.user_agent or '',
            ])

        return response

    elif export_format == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '审计日志'

            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC'),
            )

            headers = ['操作时间', '操作用户', '操作类型', '模块', '操作对象', '操作详情', 'IP地址', '用户代理']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            action_display = dict(AuditLog.ACTION_CHOICES)

            for row, log in enumerate(audit_logs, 2):
                ws.cell(row=row, column=1, value=log.created_at.strftime('%Y-%m-%d %H:%M:%S')).border = thin_border
                ws.cell(row=row, column=2, value=log.user.username if log.user else '匿名').border = thin_border
                ws.cell(row=row, column=3, value=action_display.get(log.action, '其他操作')).border = thin_border
                ws.cell(row=row, column=4, value=log.module).border = thin_border
                ws.cell(row=row, column=5, value=log.target or '').border = thin_border
                ws.cell(row=row, column=6, value=log.detail or '').border = thin_border
                ws.cell(row=row, column=7, value=log.ip_address or '').border = thin_border
                ws.cell(row=row, column=8, value=log.user_agent or '').border = thin_border

            column_widths = [20, 15, 12, 15, 25, 40, 18, 30]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="audit_logs.xlsx"'
            return response

        except ImportError:
            response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
            response['Content-Disposition'] = 'attachment; filename="audit_logs.csv"'
            response.write('\ufeff')

            writer = csv.writer(response)
            writer.writerow(['操作时间', '操作用户', '操作类型', '模块', '操作对象', '操作详情', 'IP地址', '用户代理'])

            action_display = dict(AuditLog.ACTION_CHOICES)

            for log in audit_logs:
                writer.writerow([
                    log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    log.user.username if log.user else '匿名',
                    action_display.get(log.action, '其他操作'),
                    log.module,
                    log.target or '',
                    log.detail or '',
                    log.ip_address or '',
                    log.user_agent or '',
                ])

            return response

    return HttpResponse('不支持的导出格式', status=400)
