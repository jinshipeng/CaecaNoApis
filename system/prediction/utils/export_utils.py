import io
import zipfile
import pandas as pd
from django.http import HttpResponse
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False


class DataExporter:
    """数据导出工具类"""

    @staticmethod
    def export_to_excel(data, filename=None, sheet_name='Sheet1'):
        """导出数据到Excel"""
        if filename is None:
            filename = f'export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        output = io.BytesIO()
        df = pd.DataFrame(data)
        df.to_excel(output, index=False, sheet_name=sheet_name, engine='openpyxl')
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

    @staticmethod
    def export_orders_to_excel(orders):
        """导出订单数据到Excel"""
        data = []
        for order in orders:
            data.append({
                '订单编号': order.order_no,
                '客户名称': order.customer_name,
                '物料编码': order.material.material_code if order.material else '',
                '物料名称': order.material.material_name if order.material else '',
                '订单数量': order.quantity,
                '需求日期': order.demand_date.strftime('%Y-%m-%d'),
                '优先级': order.priority,
                '状态': order.get_status_display() if hasattr(order, 'get_status_display') else order.status,
                '物流方式': order.get_shipping_method_display() if hasattr(order, 'get_shipping_method_display') else order.shipping_method,
                '物流天数': order.shipping_days,
                '创建时间': order.created_at.strftime('%Y-%m-%d %H:%M:%S') if order.created_at else ''
            })
        
        return DataExporter.export_to_excel(data, 'orders.xlsx', '订单列表')

    @staticmethod
    def export_inventory_to_excel(inventories):
        """导出库存数据到Excel"""
        data = []
        for inv in inventories:
            data.append({
                '物料编码': inv.material.material_code if inv.material else '',
                '物料名称': inv.material.material_name if inv.material else '',
                '库存数量': inv.quantity,
                '库存类型': inv.get_inventory_type_display() if hasattr(inv, 'get_inventory_type_display') else inv.inventory_type,
                '仓库': inv.warehouse,
                '批次号': inv.batch_no,
                '冻结状态': '冻结' if inv.is_hold else '正常',
                '冻结截止': inv.hold_until.strftime('%Y-%m-%d') if inv.hold_until else '',
                '安全库存': inv.material.safety_stock if inv.material else 0,
                '创建时间': inv.created_at.strftime('%Y-%m-%d %H:%M:%S') if inv.created_at else ''
            })
        
        return DataExporter.export_to_excel(data, 'inventory.xlsx', '库存列表')

    @staticmethod
    def export_shortage_report_to_excel(shortage_data):
        """导出缺料报表到Excel"""
        return DataExporter.export_to_excel(shortage_data, 'shortage_report.xlsx', '缺料报表')

    @staticmethod
    def export_to_pdf(data, title='报表', filename=None):
        if not HAS_REPORTLAB:
            return DataExporter.export_to_excel(data, filename=filename or 'export.xlsx', sheet_name=title)
        if filename is None:
            filename = f'export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=1,
            spaceAfter=20
        )
        
        elements = []
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', styles['Normal']))
        elements.append(Spacer(1, 0.25*inch))
        
        if data:
            headers = list(data[0].keys())
            table_data = [headers]
            for row in data:
                table_data.append([str(row.get(h, '')) for h in headers])
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            elements.append(table)
        
        doc.build(elements)
        
        response = HttpResponse(output.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

    @staticmethod
    def export_planning_summary_to_pdf(summary):
        """导出物料计划汇总到PDF"""
        data = [
            {'指标': '总订单数', '数值': summary.get('total_orders', 0)},
            {'指标': '齐套订单数', '数值': summary.get('complete_orders', 0)},
            {'指标': '部分齐套订单数', '数值': summary.get('partial_orders', 0)},
            {'指标': '未齐套订单数', '数值': summary.get('pending_orders', 0)},
            {'指标': '平均齐套率', '数值': f"{summary.get('avg_complete_rate', 0):.1%}"},
            {'指标': '齐套率', '数值': f"{summary.get('complete_rate', 0):.1%}"},
            {'指标': '缺料订单数', '数值': summary.get('total_shortage_orders', 0)},
            {'指标': '交期变更次数', '数值': summary.get('total_promise_changes', 0)},
            {'指标': '稳定订单数', '数值': summary.get('stable_orders', 0)},
        ]
        
        return DataExporter.export_to_pdf(data, '物料计划汇总报表', 'planning_summary.pdf')

    # ==================== 增强导出功能 ====================

    @staticmethod
    def _apply_excel_header_style(ws, row_num=1):
        """应用Excel表头样式：加粗 + 背景色"""
        if not HAS_OPENPYXL:
            return
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for cell in ws[row_num]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    @staticmethod
    def _apply_urgency_formatting(ws, col_idx, start_row=2, end_row=None):
        """应用紧急程度条件格式标色"""
        if not HAS_OPENPYXL or end_row is None:
            end_row = ws.max_row
        # 紧急程度颜色映射
        urgency_colors = {
            'critical': PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid'),   # 红色
            'urgent': PatternFill(start_color='FFA94D', end_color='FFA94D', fill_type='solid'),     # 橙色
            'normal': PatternFill(start_color='FFE066', end_color='FFE066', fill_type='solid'),    # 黄色
        }
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            val = str(cell.value or '').lower()
            if val in urgency_colors:
                cell.fill = urgency_colors[val]

    @staticmethod
    def _auto_adjust_column_width(ws, min_width=10, max_width=50):
        """自动调整列宽"""
        if not HAS_OPENPYXL:
            return
        for column_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    if cell.value:
                        # 中文字符宽度计算（中文字符约占2个英文字符宽度）
                        cell_len = 0
                        for ch in str(cell.value):
                            if '\u4e00' <= ch <= '\u9fff':
                                cell_len += 2
                            else:
                                cell_len += 1
                        max_length = max(max_length, cell_len)
                except Exception:
                    pass
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = adjusted_width

    @staticmethod
    def _write_sheet_with_style(ws, headers, data_rows, numeric_cols=None, urgency_col_idx=None):
        """通用写入Sheet并应用样式的辅助方法"""
        if not HAS_OPENPYXL:
            return
        numeric_cols = numeric_cols or []
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)

        # 写入数据行
        for row_idx, row_data in enumerate(data_rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                # 数值列右对齐
                if headers[col_idx - 1] in numeric_cols:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        # 应用表头样式
        DataExporter._apply_excel_header_style(ws, row_num=1)

        # 应用紧急程度格式
        if urgency_col_idx:
            DataExporter._apply_urgency_formatting(ws, urgency_col_idx, start_row=2, end_row=len(data_rows) + 1)

        # 自动调整列宽
        DataExporter._auto_adjust_column_width(ws)

    @staticmethod
    def export_detailed_shortage_report_to_excel(shortage_report_data, filename=None):
        """
        导出详细的缺料分析Excel报表，包含多个Sheet：
        - Sheet1 缺料汇总: 物料编码/名称/缺料数量/紧急程度/受影响订单数/推荐供应商/建议措施/预计到货日期
        - Sheet2 订单影响明细: 订单号/客户/产品/缺料物料/缺料数量/需求日期/是否影响交付/建议行动
        - Sheet3 根因分析: 缺料类别/占比/涉及物料数/建议长期措施
        - Sheet4 采购行动计划: 紧急程度/物料/需求数量/推荐供应商/最晚下单日期/运输方式建议/预算金额
        """
        if not HAS_OPENPYXL:
            # 降级处理：使用基础pandas导出
            fallback_data = shortage_report_data.get('summary', [])
            return DataExporter.export_to_excel(fallback_data, filename or 'detailed_shortage_report.xlsx', '缺料汇总')

        if filename is None:
            filename = f'detailed_shortage_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        wb = Workbook()

        # ========== Sheet1: 缺料汇总 ==========
        ws1 = wb.active
        ws1.title = '缺料汇总'
        summary_headers = ['物料编码', '物料名称', '缺料数量', '紧急程度', '受影响订单数',
                           '推荐供应商', '建议措施', '预计到货日期']
        summary_data = shortage_report_data.get('summary', [])
        summary_rows = [
            [
                item.get('material_code', ''),
                item.get('material_name', ''),
                item.get('shortage_qty', 0),
                item.get('urgency', 'normal'),
                item.get('affected_orders', 0),
                item.get('recommended_supplier', ''),
                item.get('suggested_action', ''),
                item.get('expected_arrival_date', '')
            ]
            for item in summary_data
        ]
        DataExporter._write_sheet_with_style(
            ws1, summary_headers, summary_rows,
            numeric_cols=['缺料数量', '受影响订单数'],
            urgency_col_idx=4  # 紧急程度列索引
        )

        # 添加数据透视表友好的汇总行
        pivot_row = len(summary_rows) + 3
        ws1.cell(row=pivot_row, column=1, value='【数据透视汇总】').font = Font(bold=True, color='4472C4')
        ws1.cell(row=pivot_row + 1, column=1, value='总缺料项数')
        ws1.cell(row=pivot_row + 1, column=2, value=len(summary_rows))
        ws1.cell(row=pivot_row + 2, column=1, value='总缺料数量')
        total_qty = sum(item.get('shortage_qty', 0) for item in summary_data)
        ws1.cell(row=pivot_row + 2, column=2, value=total_qty)
        critical_count = sum(1 for item in summary_data if item.get('urgency') == 'critical')
        ws1.cell(row=pivot_row + 3, column=1, value='紧急(Critical)项数')
        ws1.cell(row=pivot_row + 3, column=2, value=critical_count)

        # ========== Sheet2: 订单影响明细 ==========
        ws2 = wb.create_sheet(title='订单影响明细')
        order_headers = ['订单号', '客户', '产品', '缺料物料', '缺料数量',
                         '需求日期', '是否影响交付', '建议行动']
        order_data = shortage_report_data.get('order_details', [])
        order_rows = [
            [
                item.get('order_no', ''),
                item.get('customer', ''),
                item.get('product', ''),
                item.get('shortage_material', ''),
                item.get('shortage_qty', 0),
                item.get('demand_date', ''),
                '是' if item.get('affects_delivery') else '否',
                item.get('suggested_action', '')
            ]
            for item in order_data
        ]
        DataExporter._write_sheet_with_style(
            ws2, order_headers, order_rows,
            numeric_cols=['缺料数量']
        )

        # ========== Sheet3: 根因分析 ==========
        ws3 = wb.create_sheet(title='根因分析')
        root_cause_headers = ['缺料类别', '占比(%)', '涉及物料数', '建议长期措施']
        root_cause_data = shortage_report_data.get('root_cause_analysis', [])
        root_cause_rows = [
            [
                item.get('category', ''),
                item.get('percentage', 0),
                item.get('material_count', 0),
                item.get('long_term_measure', '')
            ]
            for item in root_cause_data
        ]
        DataExporter._write_sheet_with_style(
            ws3, root_cause_headers, root_cause_rows,
            numeric_cols=['占比(%)', '涉及物料数']
        )

        # ========== Sheet4: 采购行动计划 ==========
        ws4 = wb.create_sheet(title='采购行动计划')
        action_headers = ['紧急程度', '物料', '需求数量', '推荐供应商',
                          '最晚下单日期', '运输方式建议', '预算金额(元)']
        action_data = shortage_report_data.get('procurement_actions', [])
        action_rows = [
            [
                item.get('urgency', 'normal'),
                item.get('material', ''),
                item.get('required_qty', 0),
                item.get('supplier', ''),
                item.get('latest_order_date', ''),
                item.get('shipping_method', ''),
                item.get('budget_amount', 0)
            ]
            for item in action_data
        ]
        DataExporter._write_sheet_with_style(
            ws4, action_headers, action_rows,
            numeric_cols=['需求数量', '预算金额(元)'],
            urgency_col_idx=1  # 紧急程度列索引
        )

        # 输出到HttpResponse
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @staticmethod
    def export_management_summary_to_pdf(summary_data, filename=None):
        """
        导出PDF格式的管理层摘要报告，包含：
        - 封面标题 + 生成时间 + 数据截止时间
        - KPI仪表盘区域（4个核心指标）：齐套率/达成率/库存周转天数/交期准确率
        - 趋势分析区域（近7天的趋势描述文字）
        - Top 5 风险项表格
        - AI建议摘要（3-5条）
        - 页脚：系统名称 + 页码
        """
        if not HAS_REPORTLAB:
            # 降级为Excel导出
            fallback = [
                {'指标': '齐套率', '数值': f"{summary_data.get('kpi', {}).get('kit_completion_rate', 0):.1%}"},
                {'指标': '达成率', '数值': f"{summary_data.get('kpi', {}).get('achievement_rate', 0):.1%}"},
                {'指标': '库存周转天数', '数值': summary_data.get('kpi', {}).get('inventory_turnover_days', 0)},
                {'指标': '交期准确率', '数值': f"{summary_data.get('kpi', {}).get('delivery_accuracy', 0):.1%}"},
            ]
            return DataExporter.export_to_excel(fallback, filename or 'management_summary.xlsx', '管理层摘要')

        if filename is None:
            filename = f'management_summary_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            leftMargin=2*cm, rightMargin=2*cm,
            topMargin=2*cm, bottomMargin=2*cm
        )

        styles = getSampleStyleSheet()

        # 自定义样式
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=22,
            alignment=1,  # 居中
            spaceAfter=6,
            textColor=colors.HexColor('#1a365d'),
            fontName='Helvetica-Bold'
        )

        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=11,
            alignment=1,
            textColor=colors.HexColor('#666666'),
            spaceAfter=20
        )

        section_title_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2c5282'),
            spaceBefore=16,
            spaceAfter=8,
            fontName='Helvetica-Bold'
        )

        kpi_value_style = ParagraphStyle(
            'KPIValue',
            parent=styles['Normal'],
            fontSize=28,
            alignment=1,
            textColor=colors.HexColor('#c53030'),
            fontName='Helvetica-Bold',
            spaceAfter=2
        )

        kpi_label_style = ParagraphStyle(
            'KPILabel',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1,
            textColor=colors.HexColor('#718096'),
            spaceAfter=12
        )

        body_style = ParagraphStyle(
            'BodyText',
            parent=styles['Normal'],
            fontSize=10,
            leading=16,
            spaceAfter=6
        )

        elements = []

        # ===== 封面标题区 =====
        elements.append(Spacer(1, 0.5*inch))
        elements.append(Paragraph('智能供应链管理系统', title_style))
        elements.append(Paragraph('管理层决策摘要报告', ParagraphStyle(
            'SubTitle2', parent=title_style, fontSize=16, textColor=colors.HexColor('#4a5568'),
            spaceBefore=8, spaceAfter=12
        )))

        # 生成时间和数据截止时间
        gen_time = datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')
        cutoff_time = summary_data.get('data_cutoff_time', datetime.now().strftime('%Y-%m-%d %H:%M'))
        elements.append(Paragraph(f'报告生成时间: {gen_time}', subtitle_style))
        elements.append(Paragraph(f'数据截止时间: {cutoff_time}', subtitle_style))

        # 分隔线效果
        elements.append(Spacer(1, 0.15*inch))

        # ===== KPI仪表盘区域（4个核心指标）=====
        kpi_data = summary_data.get('kpi', {})
        elements.append(Paragraph('核心指标仪表盘', section_title_style))

        # KPI表格：4列布局
        kpi_items = [
            ('齐套率', f"{kpi_data.get('kit_completion_rate', 0):.1%}"),
            ('达成率', f"{kpi_data.get('achievement_rate', 0):.1%}"),
            ('库存周转天数', f"{kpi_data.get('inventory_turnover_days', 0):.1f}天"),
            ('交期准确率', f"{kpi_data.get('delivery_accuracy', 0):.1%}"),
        ]

        kpi_table_data = [[
            Paragraph(f'<font size="28" color="#c53030"><b>{val}</b></font>', styles['Normal']),
            Paragraph(f'<font size="9" color="#718096">{label}</font>', styles['Normal'])
        ] for label, val in kpi_items]

        kpi_table = Table(kpi_table_data, colWidths=[3.8*cm]*4)
        kpi_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f7fafc')),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(kpi_table)

        # ===== 趋势分析区域 =====
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph('趋势分析（近7日）', section_title_style))
        trend_text = summary_data.get('trend_analysis', '暂无趋势分析数据。')
        if isinstance(trend_text, list):
            trend_text = '\n'.join(trend_text)
        elements.append(Paragraph(trend_text.replace('\n', '<br/>'), body_style))

        # ===== Top 5 风险项表格 =====
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph('Top 5 高风险项', section_title_style))

        risk_items = summary_data.get('top_risks', [])
        if risk_items:
            risk_headers = ['排名', '风险描述', '影响范围', '风险等级', '建议措施']
            risk_table_data = [risk_headers]
            for idx, risk in enumerate(risk_items[:5], 1):
                risk_table_data.append([
                    str(idx),
                    risk.get('description', '')[:40],
                    risk.get('impact_scope', ''),
                    risk.get('risk_level', ''),
                    risk.get('mitigation', '')[:35]
                ])

            risk_table = Table(risk_table_data, colWidths=[1.2*cm, 5*cm, 2.5*cm, 2*cm, 5*cm])
            risk_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c5282')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7fafc')]),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(risk_table)
        else:
            elements.append(Paragraph('暂无高风险项数据。', body_style))

        # ===== AI建议摘要 =====
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph('AI智能建议', section_title_style))

        ai_suggestions = summary_data.get('ai_suggestions', [])
        if ai_suggestions:
            for idx, suggestion in enumerate(ai_suggestions[:5], 1):
                sug_text = f"<b>{idx}. {suggestion.get('title', '建议')}</b><br/>"
                sug_text += f"<font size='9' color='#4a5568'>{suggestion.get('content', '')}</font>"
                elements.append(Paragraph(sug_text, body_style))
                elements.append(Spacer(1, 0.08*inch))
        else:
            elements.append(Paragraph('暂无AI建议数据。', body_style))

        # 构建PDF
        doc.build(elements)

        response = HttpResponse(output.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @staticmethod
    def export_procurement_action_plan_to_excel(action_plan, filename=None):
        """
        导出采购行动方案Excel（来自AI引擎的 generate_procurement_action_plan 结果）：
        - 立即行动项（0-3天）
        - 短期计划（3-14天）
        - 中期规划（14-30天）
        - 优化建议
        - 风险缓解计划
        - 总投资估算
        """
        if not HAS_OPENPYXL:
            fallback = action_plan.get('immediate_actions', [])
            return DataExporter.export_to_excel(fallback, filename or 'procurement_action_plan.xlsx', '立即行动项')

        if filename is None:
            filename = f'procurement_action_plan_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        wb = Workbook()

        # ========== Sheet1: 立即行动项（0-3天）==========
        ws1 = wb.active
        ws1.title = '立即行动项(0-3天)'
        imm_headers = ['序号', '物料编码', '物料名称', '需求数量', '缺口数量',
                       '推荐供应商', '联系人', '联系电话', '最晚下单日期',
                       '运输方式', '预计到货日', '预算金额(元)', '优先级', '备注']
        imm_data = action_plan.get('immediate_actions', [])
        imm_rows = [
            [
                idx + 1,
                item.get('material_code', ''),
                item.get('material_name', ''),
                item.get('required_qty', 0),
                item.get('gap_qty', 0),
                item.get('supplier', ''),
                item.get('contact_person', ''),
                item.get('contact_phone', ''),
                item.get('latest_order_date', ''),
                item.get('shipping_method', ''),
                item.get('expected_arrival', ''),
                item.get('budget', 0),
                item.get('priority', 'high'),
                item.get('remark', '')
            ]
            for idx, item in enumerate(imm_data)
        ]
        DataExporter._write_sheet_with_style(
            ws1, imm_headers, imm_rows,
            numeric_cols=['需求数量', '缺口数量', '预算金额(元)'],
            urgency_col_idx=13  # 优先级列
        )

        # ========== Sheet2: 短期计划（3-14天）==========
        ws2 = wb.create_sheet(title='短期计划(3-14天)')
        short_headers = ['序号', '物料编码', '物料名称', '需求数量', '推荐供应商',
                         '计划下单日期', '期望到货日期', '预算金额(元)', '状态', '备注']
        short_data = action_plan.get('short_term_plan', [])
        short_rows = [
            [
                idx + 1,
                item.get('material_code', ''),
                item.get('material_name', ''),
                item.get('required_qty', 0),
                item.get('supplier', ''),
                item.get('plan_order_date', ''),
                item.get('expected_arrival', ''),
                item.get('budget', 0),
                item.get('status', '待执行'),
                item.get('remark', '')
            ]
            for idx, item in enumerate(short_data)
        ]
        DataExporter._write_sheet_with_style(
            ws2, short_headers, short_rows,
            numeric_cols=['需求数量', '预算金额(元)']
        )

        # ========== Sheet3: 中期规划（14-30天）==========
        ws3 = wb.create_sheet(title='中期规划(14-30天)')
        mid_headers = ['序号', '物料类别', '涉及物料数', '总需求数量', '预估金额(元)',
                       '建议采购策略', '目标供应商', '计划启动日期', '备注']
        mid_data = action_plan.get('mid_term_plan', [])
        mid_rows = [
            [
                idx + 1,
                item.get('category', ''),
                item.get('material_count', 0),
                item.get('total_qty', 0),
                item.get('estimated_amount', 0),
                item.get('strategy', ''),
                item.get('target_supplier', ''),
                item.get('start_date', ''),
                item.get('remark', '')
            ]
            for idx, item in enumerate(mid_data)
        ]
        DataExporter._write_sheet_with_style(
            ws3, mid_headers, mid_rows,
            numeric_cols=['涉及物料数', '总需求数量', '预估金额(元)']
        )

        # ========== Sheet4: 优化建议 ==========
        ws4 = wb.create_sheet(title='优化建议')
        opt_headers = ['序号', '建议类型', '建议内容', '预期收益', '实施难度', '优先级', '责任部门']
        opt_data = action_plan.get('optimization_suggestions', [])
        opt_rows = [
            [
                idx + 1,
                item.get('type', ''),
                item.get('content', ''),
                item.get('expected_benefit', ''),
                item.get('difficulty', ''),
                item.get('priority', ''),
                item.get('responsible_dept', '')
            ]
            for idx, item in enumerate(opt_data)
        ]
        DataExporter._write_sheet_with_style(ws4, opt_headers, opt_rows)

        # ========== Sheet5: 风险缓解计划 ==========
        ws5 = wb.create_sheet(title='风险缓解计划')
        risk_headers = ['序号', '风险描述', '发生概率', '影响程度', '风险等级',
                        '缓解措施', '责任人', '截止日期', '当前状态']
        risk_data = action_plan.get('risk_mitigation', [])
        risk_rows = [
            [
                idx + 1,
                item.get('description', ''),
                item.get('probability', ''),
                item.get('impact', ''),
                item.get('level', ''),
                item.get('mitigation_action', ''),
                item.get('owner', ''),
                item.get('deadline', ''),
                item.get('status', '待处理')
            ]
            for idx, item in enumerate(risk_data)
        ]
        DataExporter._write_sheet_with_style(
            ws5, risk_headers, risk_rows,
            urgency_col_idx=4  # 风险等级列
        )

        # ========== Sheet6: 总投资估算 ==========
        ws6 = wb.create_sheet(title='总投资估算')
        invest_headers = ['项目分类', '金额(元)', '占比(%)', '说明']
        investment = action_plan.get('investment_summary', {})
        total_amount = investment.get('total_amount', 0)
        invest_rows = [
            ['立即行动项', investment.get('immediate_amount', 0),
             f"{investment.get('immediate_amount', 0) / max(total_amount, 1) * 100:.1f}" if total_amount else '0',
             '0-3天内需执行的紧急采购'],
            ['短期计划', investment.get('short_term_amount', 0),
             f"{investment.get('short_term_amount', 0) / max(total_amount, 1) * 100:.1f}" if total_amount else '0',
             '3-14天内的采购计划'],
            ['中期规划', investment.get('mid_term_amount', 0),
             f"{investment.get('mid_term_amount', 0) / max(total_amount, 1) * 100:.1f}" if total_amount else '0',
             '14-30天的中期采购'],
            ['风险预留金', investment.get('risk_reserve', 0),
             f"{investment.get('risk_reserve', 0) / max(total_amount, 1) * 100:.1f}" if total_amount else '0',
             '应对不确定性的风险缓冲'],
            ['总计', total_amount, '100%', '全部投资估算合计'],
        ]
        DataExporter._write_sheet_with_style(
            ws6, invest_headers, invest_rows,
            numeric_cols=['金额(元)', '占比(%)']
        )
        # 总计行加粗高亮
        if HAS_OPENPYXL and len(invest_rows) >= 5:
            last_row = len(invest_rows) + 1
            for col in range(1, 5):
                cell = ws6.cell(row=last_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

        # 输出
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @staticmethod
    def export_full_analysis_package(report_data, format='zip'):
        """
        一键导出完整分析包：调用上述所有导出方法，打包成ZIP
        包含：
        - detailed_shortage_report.xlsx（详细缺料报表）
        - management_summary.pdf（管理层摘要PDF）
        - procurement_action_plan.xlsx（采购行动方案）
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        zip_filename = f'full_analysis_package_{timestamp}.zip'

        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            # 1. 导出详细缺料报表Excel
            try:
                shortage_response = DataExporter.export_detailed_shortage_report_to_excel(
                    report_data.get('shortage_report', {}),
                    filename=f'detailed_shortage_report_{timestamp}.xlsx'
                )
                zf.writestr(f'detailed_shortage_report_{timestamp}.xlsx', shortage_response.content)
            except Exception as e:
                # 写入错误信息文件
                zf.writestr('ERROR_detailed_shortage_report.txt', f'缺料报表导出失败: {str(e)}')

            # 2. 导出管理层摘要PDF
            try:
                pdf_response = DataExporter.export_management_summary_to_pdf(
                    report_data.get('summary_data', {}),
                    filename=f'management_summary_{timestamp}.pdf'
                )
                zf.writestr(f'management_summary_{timestamp}.pdf', pdf_response.content)
            except Exception as e:
                zf.writestr('ERROR_management_summary.txt', f'管理层摘要导出失败: {str(e)}')

            # 3. 导出采购行动方案Excel
            try:
                proc_response = DataExporter.export_procurement_action_plan_to_excel(
                    report_data.get('action_plan', {}),
                    filename=f'procurement_action_plan_{timestamp}.xlsx'
                )
                zf.writestr(f'procurement_action_plan_{timestamp}.xlsx', proc_response.content)
            except Exception as e:
                zf.writestr('ERROR_procurement_action_plan.txt', f'采购方案导出失败: {str(e)}')

            # 4. 添加README说明文件
            readme_content = f"""智能供应链分析报告包
========================
生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
系统版本: 智能供应链预测系统 v1.0

文件清单:
---------
1. detailed_shortage_report_{timestamp}.xlsx
   - 缺料汇总（含条件格式和透视汇总）
   - 订单影响明细
   - 根因分析
   - 采购行动计划

2. management_summary_{timestamp}.pdf
   - 管理层决策摘要
   - KPI仪表盘
   - 趋势分析与Top5风险
   - AI智能建议

3. procurement_action_plan_{timestamp}.xlsx
   - 立即行动项(0-3天)
   - 短期计划(3-14天)
   - 中期规划(14-30天)
   - 优化建议与风险缓解
   - 总投资估算

注意: 本报告由AI引擎自动生成，仅供参考。
"""
            zf.writestr('README.txt', readme_content)

        zip_buffer.seek(0)

        response = HttpResponse(
            zip_buffer.getvalue(),
            content_type='application/zip'
        )
        response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
        return response