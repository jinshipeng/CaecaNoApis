import csv
import io
from datetime import datetime
from django.http import HttpResponse
from django.db.models import QuerySet


class DataExporter:
    """数据导出工具类 - 支持多种格式"""

    # 安全限制：最大导出行数（防止内存溢出）
    MAX_EXPORT_ROWS = 10000

    @staticmethod
    def export_to_csv(queryset, filename='data', fields=None, headers=None):
        """
        导出数据为CSV格式

        Args:
            queryset: Django QuerySet或数据列表
            filename: 文件名（不含扩展名）
            fields: 要导出的字段列表（如果是QuerySet）
            headers: 自定义表头

        Returns:
            HttpResponse: CSV文件响应
        """
        # 安全检查：限制导出行数
        if isinstance(queryset, (list, tuple)) and len(queryset) > DataExporter.MAX_EXPORT_ROWS:
            queryset = queryset[:DataExporter.MAX_EXPORT_ROWS]
        elif hasattr(queryset, 'count') and queryset.count() > DataExporter.MAX_EXPORT_ROWS:
            queryset = queryset[:DataExporter.MAX_EXPORT_ROWS]

        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        
        # 处理QuerySet
        if isinstance(queryset, QuerySet):
            model = queryset.model
            if not fields:
                fields = [field.name for field in model._meta.fields]
            
            if not headers:
                headers = [field.verbose_name or field.name for field in model._meta.fields 
                          if field.name in fields]
            
            writer.writerow(headers)
            
            for obj in queryset:
                row = []
                for field in fields:
                    value = getattr(obj, field, '')
                    if value is None:
                        value = ''
                    elif hasattr(value, '__str__'):
                        value = str(value)
                    row.append(value)
                writer.writerow(row)
        
        # 处理普通列表/字典列表
        else:
            if isinstance(queryset, list) and len(queryset) > 0:
                if isinstance(queryset[0], dict):
                    if not fields:
                        fields = list(queryset[0].keys())
                    if not headers:
                        headers = fields
                    
                    writer.writerow(headers)
                    
                    for item in queryset:
                        row = [item.get(field, '') for field in fields]
                        writer.writerow(row)
                
                elif isinstance(queryset[0], (list, tuple)):
                    if not headers:
                        headers = [f'Column_{i+1}' for i in range(len(queryset[0]))]
                    writer.writerow(headers)
                    writer.writerows(queryset)
        
        return response
    
    @staticmethod
    def export_to_excel(queryset, filename='data', sheet_name='Sheet1', fields=None):
        """
        导出数据为Excel格式（使用openpyxl）

        Args:
            queryset: 数据源
            filename: 文件名
            sheet_name: 工作表名称
            fields: 字段列表

        Returns:
            HttpResponse: Excel文件响应
        """
        # 安全检查：限制导出行数
        if isinstance(queryset, (list, tuple)) and len(queryset) > DataExporter.MAX_EXPORT_ROWS:
            queryset = queryset[:DataExporter.MAX_EXPORT_ROWS]
        elif hasattr(queryset, 'count') and queryset.count() > DataExporter.MAX_EXPORT_ROWS:
            queryset = queryset[:DataExporter.MAX_EXPORT_ROWS]

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            return DataExporter.export_to_csv(queryset, filename, fields)
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 定义样式
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        data_rows = []
        
        # 处理QuerySet
        if isinstance(queryset, QuerySet):
            model = queryset.model
            if not fields:
                fields = [field.name for field in model._meta.fields]
            
            headers = [field.verbose_name or field.name for field in model._meta.fields 
                      if field.name in fields]
            
            for obj in queryset:
                row = []
                for field in fields:
                    value = getattr(obj, field, '')
                    if value is None:
                        value = ''
                    elif hasattr(value, '__str__'):
                        value = str(value)
                    row.append(value)
                data_rows.append(row)
        
        # 处理字典列表
        elif isinstance(queryset, list) and len(queryset) > 0 and isinstance(queryset[0], dict):
            if not fields:
                fields = list(queryset[0].keys())
            headers = fields
            
            for item in queryset:
                row = [item.get(field, '') for field in fields]
                data_rows.append(row)
        
        else:
            return DataExporter.export_to_csv(queryset, filename, fields)
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 写入数据
        for row_num, row_data in enumerate(data_rows, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
        
        # 调整列宽
        for col_num in range(1, len(headers) + 1):
            column_letter = ws.cell(row=1, column=col_num).column_letter
            max_length = max(
                len(str(headers[col_num - 1])),
                *[len(str(row[col_num - 1])) for row in data_rows]
            ) if data_rows else len(str(headers[col_num - 1]))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
    
    @staticmethod
    def export_to_json(queryset, filename='data', fields=None):
        """
        导出数据为JSON格式
        
        Args:
            queryset: 数据源
            filename: 文件名
            fields: 字段列表
            
        Returns:
            HttpResponse: JSON文件响应
        """
        import json
        
        data = []
        
        if isinstance(queryset, QuerySet):
            if not fields:
                fields = [field.name for field in queryset.model._meta.fields]
            
            for obj in queryset:
                item = {}
                for field in fields:
                    value = getattr(obj, field, None)
                    if value is not None:
                        if hasattr(value, 'isoformat'):
                            value = value.isoformat()
                        elif hasattr(value, '__str__'):
                            value = str(value)
                    item[field] = value
                data.append(item)
        
        elif isinstance(queryset, list):
            if fields:
                for item in queryset:
                    if isinstance(item, dict):
                        filtered_item = {k: v for k, v in item.items() if k in fields}
                        data.append(filtered_item)
                    else:
                        data.append(item)
            else:
                data = queryset
        
        response = HttpResponse(
            json.dumps(data, ensure_ascii=False, indent=2),
            content_type='application/json; charset=utf-8'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json"'
        
        return response


class BatchOperationManager:
    """批量操作管理器"""
    
    @staticmethod
    def batch_delete(model_class, ids, user=None):
        """
        批量删除记录
        
        Args:
            model_class: Django模型类
            ids: 要删除的ID列表
            user: 操作用户（用于日志）
            
        Returns:
            (success_count, error_count, errors) 元组
        """
        success_count = 0
        error_count = 0
        errors = []
        
        try:
            queryset = model_class.objects.filter(id__in=ids)
            count, _ = queryset.delete()
            success_count = count
            
        except Exception as e:
            error_count = len(ids)
            errors.append(f'批量删除失败: {str(e)}')
        
        return success_count, error_count, errors
    
    @staticmethod
    def batch_update(model_class, ids, update_data, user=None):
        """
        批量更新记录
        
        Args:
            model_class: Django模型类
            ids: 要更新的ID列表
            update_data: 更新的数据字典 {field: value}
            user: 操作用户
            
        Returns:
            (success_count, error_count, errors) 元组
        """
        success_count = 0
        error_count = 0
        errors = []
        
        try:
            queryset = model_class.objects.filter(id__in=ids)
            updated = queryset.update(**update_data)
            success_count = updated
            
        except Exception as e:
            error_count = len(ids)
            errors.append(f'批量更新失败: {str(e)}')
        
        return success_count, error_count, errors
    
    @staticmethod
    def batch_status_change(model_class, ids, status_field, new_status, user=None):
        """
        批量更改状态
        
        Args:
            model_class: Django模型类
            ids: 要更改的ID列表
            status_field: 状态字段名
            new_status: 新状态值
            user: 操作用户
            
        Returns:
            (success_count, error_count, errors) 元组
        """
        return BatchOperationManager.batch_update(
            model_class, 
            ids, 
            {status_field: new_status},
            user
        )
    
    @staticmethod
    def validate_ids(ids, model_class=None):
        """
        验证ID列表的有效性
        
        Args:
            ids: ID列表
            model_class: 模型类（可选，用于验证ID是否存在）
            
        Returns:
            (is_valid, valid_ids, errors) 元组
        """
        if not ids:
            return False, [], ['未选择任何记录']
        
        if not isinstance(ids, (list, tuple)):
            ids = [ids]
        
        valid_ids = []
        errors = []
        
        for id_val in ids:
            try:
                int_id = int(id_val)
                if int_id > 0:
                    valid_ids.append(int_id)
                else:
                    errors.append(f'无效的ID: {id_val}')
            except (ValueError, TypeError):
                errors.append(f'ID必须是数字: {id_val}')
        
        # 验证ID是否存在（如果提供了model_class）
        if model_class and valid_ids:
            existing_ids = set(
                model_class.objects.filter(id__in=valid_ids).values_list('id', flat=True)
            )
            invalid_ids = set(valid_ids) - existing_ids
            
            if invalid_ids:
                errors.extend([f'ID不存在: {id}' for id in invalid_ids])
                valid_ids = list(existing_ids)
        
        is_valid = len(errors) == 0 and len(valid_ids) > 0
        
        return is_valid, valid_ids, errors


class ImportResult:
    """导入结果封装类"""
    
    def __init__(self):
        self.success_count = 0
        self.error_count = 0
        self.warning_count = 0
        self.errors = []
        self.warnings = []
        self.created_objects = []
    
    def add_success(self, obj):
        self.success_count += 1
        self.created_objects.append(obj)
    
    def add_error(self, message, row_num=None):
        self.error_count += 1
        prefix = f'第{row_num}行: ' if row_num else ''
        self.errors.append(prefix + message)
    
    def add_warning(self, message, row_num=None):
        self.warning_count += 1
        prefix = f'第{row_num}行: ' if row_num else ''
        self.warnings.append(prefix + message)
    
    def to_dict(self):
        return {
            'success': self.success_count,
            'errors': self.error_count,
            'warnings': self.warning_count,
            'error_details': self.errors,
            'warning_details': self.warnings,
            'total': self.success_count + self.error_count
        }
    
    def __repr__(self):
        return (
            f'ImportResult(success={self.success_count}, '
            f'errors={self.error_count}, warnings={self.warning_count})'
        )


def get_export_format_from_request(request):
    """
    从请求中获取导出格式
    
    Args:
        request: HTTP请求对象
        
    Returns:
        str: 导出格式 ('csv', 'excel', 'json')，默认返回'csv'
    """
    format_param = request.GET.get('format', request.POST.get('format', 'csv'))
    format_map = {
        'csv': 'csv',
        'xlsx': 'excel',
        'xls': 'excel',
        'excel': 'excel',
        'json': 'json',
    }
    return format_map.get(format_param.lower(), 'csv')


def prepare_export_response(queryset, format_type, filename='data', fields=None):
    """
    根据格式类型准备导出响应
    
    Args:
        queryset: 数据源
        format_type: 格式类型 ('csv', 'excel', 'json')
        filename: 文件名
        fields: 字段列表
        
    Returns:
        HttpResponse: 文件下载响应
    """
    exporter = DataExporter()
    
    if format_type == 'excel':
        return exporter.export_to_excel(queryset, filename, fields=fields)
    elif format_type == 'json':
        return exporter.export_to_json(queryset, filename, fields=fields)
    else:
        return exporter.export_to_csv(queryset, filename, fields=fields)